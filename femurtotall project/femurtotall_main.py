#femur length를 이용하여 tall을 추정
#140번줄에서 저장위치 변경

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

wb = Workbook()
ws = wb.active
ws.title = '결과'

count = 2

#style
r1c1 = ws.cell(row = 1, column = 1)
r1c1.value = "번호"
r1c1.font = Font(bold=True)
r1c1.alignment = Alignment(horizontal='center')
r1c2 = ws.cell(row = 1, column = 2)
r1c2.value = "성별"
r1c2.font = Font(bold=True)
r1c2.alignment = Alignment(horizontal='center')
r1c3 = ws.cell(row = 1, column = 3)
r1c3.value = "길이"
r1c3.font = Font(bold=True)
r1c3.alignment = Alignment(horizontal='center')
r1c4 = ws.cell(row = 1, column = 4)
r1c4.value = "Pearson식"
r1c4.font = Font(bold=True)
r1c4.alignment = Alignment(horizontal='center')
r1c5 = ws.cell(row = 1, column = 5)
r1c5.value = "藤井식"
r1c5.font = Font(bold=True)
r1c5.alignment = Alignment(horizontal='center')
r1c6 = ws.cell(row = 1, column = 6)
r1c6.value = "Trotter&Gleser식"
r1c6.font = Font(bold=True)
r1c6.alignment = Alignment(horizontal='center')

print("="*30)
print("="*13 + "tall" + "="*13)
print("="*30)
print(" ")

while True :
    print("남성 : 1\n여성 : 2\n종료 : 3")
    print(" ")
    MorF = input(">>> ")
    print(" ")

    if MorF == '1' : #남성
        print("남성을 입력받았습니다.")
        print("Pearson식, Trotter&Glaser식, 藤井식을 사용하여 신장을 추정합니다.")
        print("'0'을 입력하면 초기 메뉴로 돌아갑니다.")
        print(" ")
        while True :
            femur = float(input("대퇴골의 길이를 입력해주세요 >>> "))
            if femur != 0 :
                pearson_result = 81.306 + 1.880 * femur
                huzii_result = (2.47 * (femur*10) + 549.01)/10
                trotter_result = 2.15 * femur + 72.57

                soonseo = ws.cell(row = count, column = 1)
                soonseo.value = count - 1
                soonseo.alignment = Alignment(horizontal='center')
                sex = ws.cell(row = count, column = 2)
                sex.value = "남성"
                sex.alignment = Alignment(horizontal='center')
                femur_length = ws.cell(row = count, column = 3)
                femur_length.value = femur
                femur_length.alignment = Alignment(horizontal='center')
                pearsoncell = ws.cell(row = count, column = 4)
                pearsoncell.alignment = Alignment(horizontal='center')
                huziicell = ws.cell(row = count, column = 5)
                huziicell.alignment = Alignment(horizontal='center')
                trottercell = ws.cell(row = count, column = 6)
                trottercell.alignment = Alignment(horizontal='center')

                pearsoncell.value = pearson_result
                huziicell.value = huzii_result
                trottercell.value = trotter_result

                count += 1

                print("- Pearson식 : %0.3fcm" % pearson_result)
                print("- 藤井식 : %0.3fcm" % huzii_result)
                print("- Trotter&Glaser식 : %0.3fcm" % trotter_result)
                print(" ")

            elif femur == 0 :
                print(" ")
                print("=" * 30)
                print("=" * 4 + "초기 메뉴로 돌아갑니다" + "=" * 4)
                print("=" * 30)
                print(" ")
                break
            
    elif MorF == '2' : #여성
        print("여성을 입력받았습니다.")
        print("Pearson식, 藤井식을 사용하여 신장을 추정합니다.")
        print("'0'을 입력하면 초기 메뉴로 돌아갑니다.")
        print(" ")
        while True :
            femur = float(input("대퇴골의 길이를 입력해주세요 >>> "))
            if femur != 0 :
                pearson_result = 72.844 + 1.945 * femur
                huzii_result = (2.24 * (femur*10) + 610.43)/10

                soonseo = ws.cell(row = count, column = 1)
                soonseo.value = count - 1
                soonseo.alignment = Alignment(horizontal='center')
                sex = ws.cell(row = count, column = 2)
                sex.value = "여성"
                sex.alignment = Alignment(horizontal='center')
                femur_length = ws.cell(row = count, column = 3)
                femur_length.value = femur
                femur_length.alignment = Alignment(horizontal='center')
                pearsoncell = ws.cell(row = count, column = 4)
                pearsoncell.alignment = Alignment(horizontal='center')
                huziicell = ws.cell(row = count, column = 5)
                huziicell.alignment = Alignment(horizontal='center')

                pearsoncell.value = pearson_result
                huziicell.value = huzii_result

                count += 1

                print("- Pearson식 : %0.3fcm"  % pearson_result)
                print("- 藤井식 : %0.3fcm" % huzii_result)
                print(" ")

            elif femur == 0 :
                print(" ")
                print("=" * 30)
                print("=" * 4 + "초기 메뉴로 돌아갑니다" + "=" * 4)
                print("=" * 30)
                print(" ")
                break

    elif MorF == '3' : #종료
        wb.save('/Users/jch/Desktop/result.xlsx')
        print("="*30)
        print("="*10 + "종료합니다" + "="*10)
        print("="*30)
        print(" ")
        break
    
    else : #오류
        print(" ")
        print("오류입니다. 다시 입력해주세요.")
        print(" ")